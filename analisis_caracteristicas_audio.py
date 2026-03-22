#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import fnmatch
import json
import re
import subprocess
from pathlib import Path
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


AUDIO_EXTENSIONS = {
    ".mp3", ".flac", ".wav", ".m4a", ".aac", ".ogg", ".oga", ".opus",
    ".wma", ".aiff", ".aif", ".ape", ".mp4", ".3gp", ".3gpp"
}


def run_ffprobe(file_path: Path) -> dict:
    cmd = [
        "ffprobe",
        "-v", "quiet",
        "-print_format", "json",
        "-show_format",
        "-show_streams",
        str(file_path)
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8")
    if result.returncode != 0:
        raise RuntimeError(result.stderr.strip() or "ffprobe no pudo analizar el archivo")
    return json.loads(result.stdout)


def format_duration(seconds) -> str:
    try:
        seconds = float(seconds)
    except (TypeError, ValueError):
        return "No disponible"

    total_seconds = int(round(seconds))
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    secs = total_seconds % 60
    return f"{hours:02d}:{minutes:02d}:{secs:02d}"


def format_bitrate(bit_rate) -> str:
    try:
        bit_rate = int(float(bit_rate))
        return f"{round(bit_rate / 1000)} kbps"
    except (TypeError, ValueError):
        return "No disponible"


def format_sample_rate_khz(sample_rate) -> str:
    try:
        sr = int(sample_rate)
        khz = sr / 1000
        if khz.is_integer():
            return f"{int(khz)} kHz"
        return f"{khz:.1f} kHz"
    except (TypeError, ValueError):
        return "No disponible"


def format_channels(channels) -> str:
    try:
        channels = int(channels)
    except (TypeError, ValueError):
        return "No disponible"

    if channels == 1:
        return "MONO"
    if channels == 2:
        return "STEREO"
    return f"{channels} canales"


def normalize_utc_datetime(text: str) -> str:
    if not text:
        return "No disponible"

    text = str(text).strip().replace("\x00", "")
    text = text.replace("Z", "+00:00")

    try:
        dt = datetime.fromisoformat(text)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        else:
            dt = dt.astimezone(timezone.utc)
        return dt.strftime("%Y-%m-%d %H:%M:%S UTC")
    except ValueError:
        pass

    candidates = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d",
        "%Y%m%dT%H%M%S",
        "%Y%m%d",
    ]

    for fmt in candidates:
        try:
            dt = datetime.strptime(text, fmt)
            dt = dt.replace(tzinfo=timezone.utc)
            if "H" in fmt:
                return dt.strftime("%Y-%m-%d %H:%M:%S UTC")
            return dt.strftime("%Y-%m-%d 00:00:00 UTC")
        except ValueError:
            pass

    return text


def get_encoding_date_utc(data: dict) -> str:
    preferred_keys = [
        "creation_time",
        "encoded_date",
        "com.apple.quicktime.creationdate",
        "date",
        "tagged_date",
        "ICRD",
        "year",
    ]

    format_tags = data.get("format", {}).get("tags", {}) or {}
    for key in preferred_keys:
        if key in format_tags and format_tags[key]:
            return normalize_utc_datetime(format_tags[key])

    for stream in data.get("streams", []):
        tags = stream.get("tags", {}) or {}
        for key in preferred_keys:
            if key in tags and tags[key]:
                return normalize_utc_datetime(tags[key])

    return "No disponible"


def friendly_codec_name(stream: dict) -> str:
    codec_name = (stream.get("codec_name") or "").lower()
    profile = (stream.get("profile") or "").strip()
    codec_tag_string = (stream.get("codec_tag_string") or "").strip()

    if codec_name == "aac":
        if profile:
            return f"AAC {profile}"
        return "AAC"

    if codec_name == "mp3":
        return "MP3"

    if codec_name in ("pcm_s16le", "pcm_s24le", "pcm_u8", "pcm_alaw", "pcm_mulaw"):
        return codec_name.upper()

    if codec_name == "opus":
        return "Opus"

    if codec_name == "vorbis":
        return "Vorbis"

    if codec_name == "flac":
        return "FLAC"

    if codec_name == "alac":
        return "ALAC"

    if codec_name:
        if profile:
            return f"{codec_name.upper()} {profile}"
        return codec_name.upper()

    if codec_tag_string:
        return codec_tag_string

    return "No disponible"

import tempfile  # Añadir esta importación al principio del script


def parse_rms_with_ffmpeg(file_path: Path):
    """
    Versión ultra-robusta optimizada para audios periciales de larga duración.
    Evita el bloqueo de tuberías (Pipes) mediante el descarte de salida estándar.
    """
    cmd = [
        "ffmpeg",
        "-nostdin",
        "-hide_banner",
        "-v", "error", # Solo errores críticos para no saturar el log
        "-i", str(file_path),
        "-af", "astats=metadata=1:reset=0",
        "-f", "null",
        "-"
    ]

    try:
        # Ejecutamos con captura controlada solo del error (donde viajan los stats)
        # Usamos un tiempo de espera amplio pero finito
        result = subprocess.run(
            cmd,
            capture_output=True, # Solo capturamos al final, no en tiempo real
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=600 # 10 minutos máximo para audios muy pesados
        )
        text = result.stderr # FFmpeg vuelca las estadísticas en stderr
    except subprocess.TimeoutExpired:
        return "Timeout (Audio muy largo)", "", ""
    except Exception as e:
        return f"Error: {str(e)}", "", ""

    # Búsqueda de valores RMS (se mantiene tu lógica original)
    channel_matches = re.findall(
        r"Channel:\s*(\d+).*?RMS level dB:\s*([-\d\.]+|inf|-inf)",
        text,
        flags=re.DOTALL | re.IGNORECASE
    )

    def clean_rms(value: str) -> str:
        v = value.strip()
        if v.lower() in {"-inf", "inf"}: return v
        try: return f"{float(v):.2f} dB"
        except ValueError: return "No disponible"

    if channel_matches:
        if len(channel_matches) == 1:
            return clean_rms(channel_matches[0][1]), "", ""
        if len(channel_matches) >= 2:
            rms_l = clean_rms(channel_matches[0][1])
            rms_r = clean_rms(channel_matches[1][1])
            return "", rms_l, rms_r

    return "No disponible", "No disponible", "No disponible"







def analyze_audio_file(file_path: Path):
    try:
        data = run_ffprobe(file_path)

        audio_stream = None
        for stream in data.get("streams", []):
            if stream.get("codec_type") == "audio":
                audio_stream = stream
                break

        if not audio_stream:
            return (
                file_path.name,
                "No disponible",
                "No disponible",
                "No disponible",
                "No disponible",
                "No disponible",
                "No disponible",
                "No disponible",
                "No disponible",
                "No disponible",
            )

        fmt = data.get("format", {}) or {}
        channels_num = audio_stream.get("channels")

        duration = format_duration(
            audio_stream.get("duration") or fmt.get("duration")
        )

        bitrate = format_bitrate(
            audio_stream.get("bit_rate") or fmt.get("bit_rate")
        )

        sample_rate = format_sample_rate_khz(audio_stream.get("sample_rate"))
        channels = format_channels(channels_num)
        codec = friendly_codec_name(audio_stream)
        encoding_date_utc = get_encoding_date_utc(data)

        rms_mono, rms_l, rms_r = parse_rms_with_ffmpeg(file_path)

        # Ajuste final según número de canales
        try:
            channels_num = int(channels_num)
        except (TypeError, ValueError):
            channels_num = None

        if channels_num == 1:
            rms_l = ""
            rms_r = ""
        elif channels_num == 2:
            rms_mono = ""
        else:
            rms_mono = ""
            rms_l = ""
            rms_r = ""

        return (
            file_path.name,
            encoding_date_utc,
            duration,
            sample_rate,
            channels,
            bitrate,
            codec,
            rms_mono,
            rms_l,
            rms_r,
        )

    except Exception as exc:
        return (
            file_path.name,
            f"Error: {exc}",
            "No disponible",
            "No disponible",
            "No disponible",
            "No disponible",
            "No disponible",
            "No disponible",
            "No disponible",
            "No disponible",
        )


def find_matching_audio_files(folder: Path, pattern: str, recursive: bool = False):
    if recursive:
        candidates = [p for p in folder.rglob("*") if p.is_file()]
    else:
        candidates = [p for p in folder.iterdir() if p.is_file()]

    matches = []
    for path in candidates:
        suffixes = "".join(path.suffixes).lower()
        if (
            path.suffix.lower() in AUDIO_EXTENSIONS
            or suffixes.endswith(".3gpp")
            or suffixes.endswith(".3gp")
        ):
            if fnmatch.fnmatch(path.name, pattern):
                matches.append(path)

    return sorted(matches, key=lambda p: p.name.lower())


def autosize_columns(ws):
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)


def export_to_excel(rows, output_file: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Audios"

    headers = [
        "Nombre del archivo",
        "Fecha de codificación (UTC)",
        "Duración",
        "Frecuencia de muestreo",
        "Canales",
        "Bitrate",
        "Codec",
        "RMS mono",
        "RMS canal L",
        "RMS canal R",
    ]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append(row)

    autosize_columns(ws)
    wb.save(output_file)


def main():
    parser = argparse.ArgumentParser(
        description="Analiza archivos de audio y genera un Excel progresivo."
    )
    parser.add_argument("pattern", help='Patrón de nombre, por ejemplo: "*COMPLETO*"')
    parser.add_argument("-d", "--dir", default=".", help="Carpeta a analizar.")
    parser.add_argument("-o", "--output", default="reporte_audios.xlsx", help="Excel de salida.")
    parser.add_argument("-r", "--recursive", action="store_true", help="Buscar en subcarpetas.")
    args = parser.parse_args()

    folder = Path(args.dir).resolve()
    output_file = Path(args.output).resolve()

    files = find_matching_audio_files(folder, args.pattern, recursive=args.recursive)

    if not files:
        print(f"No se encontraron archivos con el patrón {args.pattern}")
        return

    print(f"Se han encontrado {len(files)} archivo(s). Analizando...")
    
    rows = []
    for i, file_path in enumerate(files, 1):
        print(f"[{i}/{len(files)}] Procesando: {file_path.name}...", end="\r")
        
        # Analizamos el archivo
        resultado = analyze_audio_file(file_path)
        rows.append(resultado)
        
        # GUARDADO PROGRESIVO: Guardamos el Excel tras cada archivo analizado
        export_to_excel(rows, output_file)

    print(f"\n\nProceso finalizado con éxito.")
    print(f"Excel actualizado y cerrado en: {output_file}")

if __name__ == "__main__":
    main()